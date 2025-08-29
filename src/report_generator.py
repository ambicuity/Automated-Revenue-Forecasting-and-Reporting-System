"""
Automated Report Generator
Creates comprehensive business reports in multiple formats
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
import sys
import os
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import LineChart, Reference, BarChart

# Add config to path
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from config.settings import *

# Setup logging
logging.basicConfig(level=getattr(logging, LOG_LEVEL), format=LOG_FORMAT)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate automated business reports in various formats"""
    
    def __init__(self):
        self.processed_data_dir = PROCESSED_DATA_DIR
        self.reports_dir = REPORTS_DIR
        self.reports_dir.mkdir(exist_ok=True)
        
        # Report timestamp
        self.report_timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
    def load_processed_data(self) -> dict:
        """Load all processed data for reporting"""
        logger.info("Loading processed data for reporting...")
        
        data = {}
        
        try:
            # Load revenue data
            data['revenue'] = pd.read_csv(self.processed_data_dir / 'processed_revenue.csv')
            data['revenue']['date'] = pd.to_datetime(data['revenue']['date'])
            
            # Load KPI data
            data['monthly_kpis'] = pd.read_csv(self.processed_data_dir / 'monthly_kpis.csv')
            data['monthly_kpis']['date'] = pd.to_datetime(data['monthly_kpis']['date'])
            
            data['unit_kpis'] = pd.read_csv(self.processed_data_dir / 'unit_kpis.csv')
            data['advanced_kpis'] = pd.read_csv(self.processed_data_dir / 'advanced_kpis.csv')
            data['alerts'] = pd.read_csv(self.processed_data_dir / 'performance_alerts.csv')
            
            # Load forecasts
            data['forecasts'] = pd.read_csv(self.processed_data_dir / 'revenue_forecasts.csv')
            data['forecasts']['date'] = pd.to_datetime(data['forecasts']['date'])
            
            logger.info("Successfully loaded all processed data")
            return data
            
        except Exception as e:
            logger.error(f"Error loading processed data: {e}")
            raise
    
    def create_executive_summary(self, data: dict) -> dict:
        """Create executive summary metrics"""
        logger.info("Creating executive summary...")
        
        latest_month = data['monthly_kpis'].iloc[-1]
        latest_unit_kpis = data['unit_kpis']
        
        summary = {
            'report_date': datetime.now().strftime('%B %d, %Y'),
            'period': latest_month['date'].strftime('%B %Y'),
            'key_metrics': {
                'total_revenue': latest_month['revenue'],
                'revenue_growth_yoy': latest_month['revenue_yoy_growth'],
                'revenue_growth_mom': latest_month['revenue_mom_growth'],
                'total_customers': latest_month['customer_count'],
                'avg_revenue_per_customer': latest_month['revenue_per_customer'],
                'profit_margin': latest_month['profit_margin'],
                'marketing_roi': latest_month['marketing_roi']
            },
            'top_performing_unit': latest_unit_kpis.loc[latest_unit_kpis['current_revenue'].idxmax(), 'business_unit'],
            'fastest_growing_unit': latest_unit_kpis.loc[latest_unit_kpis['yoy_growth'].idxmax(), 'business_unit'],
            'total_forecast_revenue': data['forecasts']['ensemble_forecast'].sum(),
            'alerts_count': len(data['alerts']),
            'high_priority_alerts': len(data['alerts'][data['alerts']['severity'] == 'High'])
        }
        
        return summary
    
    def generate_pdf_report(self, data: dict, summary: dict):
        """Generate comprehensive PDF report"""
        logger.info("Generating PDF report...")
        
        filename = f"revenue_report_{self.report_timestamp}.pdf"
        filepath = self.reports_dir / filename
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(REPORT_TITLE, title_style))
        story.append(Paragraph(f"Report Date: {summary['report_date']}", styles['Normal']))
        story.append(Paragraph(f"Reporting Period: {summary['period']}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", styles['Heading2']))
        
        exec_data = [
            ['Metric', 'Value'],
            ['Total Revenue', f"${summary['key_metrics']['total_revenue']:,.0f}"],
            ['YoY Growth', f"{summary['key_metrics']['revenue_growth_yoy']:.1%}" if pd.notna(summary['key_metrics']['revenue_growth_yoy']) else "N/A"],
            ['MoM Growth', f"{summary['key_metrics']['revenue_growth_mom']:.1%}" if pd.notna(summary['key_metrics']['revenue_growth_mom']) else "N/A"],
            ['Total Customers', f"{summary['key_metrics']['total_customers']:,.0f}"],
            ['Revenue per Customer', f"${summary['key_metrics']['avg_revenue_per_customer']:,.0f}"],
            ['Profit Margin', f"{summary['key_metrics']['profit_margin']:.1%}"],
            ['Marketing ROI', f"{summary['key_metrics']['marketing_roi']:.1f}x"]
        ]
        
        exec_table = Table(exec_data)
        exec_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(exec_table)
        story.append(Spacer(1, 20))
        
        # Key Insights
        story.append(Paragraph("Key Insights", styles['Heading2']))
        story.append(Paragraph(f"• Top performing unit: {summary['top_performing_unit']}", styles['Bullet']))
        story.append(Paragraph(f"• Fastest growing unit: {summary['fastest_growing_unit']}", styles['Bullet']))
        story.append(Paragraph(f"• Forecasted revenue (next 12 months): ${summary['total_forecast_revenue']:,.0f}", styles['Bullet']))
        story.append(Paragraph(f"• Performance alerts: {summary['alerts_count']} ({summary['high_priority_alerts']} high priority)", styles['Bullet']))
        story.append(Spacer(1, 20))
        
        # Business Unit Performance
        story.append(Paragraph("Business Unit Performance", styles['Heading2']))
        
        unit_data = [['Business Unit', 'Current Revenue', 'YoY Growth', 'Profit Margin']]
        for _, unit in data['unit_kpis'].iterrows():
            unit_data.append([
                unit['business_unit'],
                f"${unit['current_revenue']:,.0f}",
                f"{unit['yoy_growth']:.1%}",
                f"{unit['profit_margin']:.1%}"
            ])
        
        unit_table = Table(unit_data)
        unit_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(unit_table)
        story.append(Spacer(1, 20))
        
        # Performance Alerts
        if not data['alerts'].empty:
            story.append(Paragraph("Performance Alerts", styles['Heading2']))
            
            alert_data = [['Type', 'Business Unit', 'Message', 'Severity']]
            for _, alert in data['alerts'].head(10).iterrows():  # Top 10 alerts
                alert_data.append([
                    alert['type'],
                    alert['business_unit'],
                    alert['message'][:50] + '...' if len(alert['message']) > 50 else alert['message'],
                    alert['severity']
                ])
            
            alert_table = Table(alert_data, colWidths=[1.5*inch, 1*inch, 2.5*inch, 1*inch])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(alert_table)
        
        # Build PDF
        doc.build(story)
        logger.info(f"PDF report saved to {filepath}")
        
    def generate_excel_report(self, data: dict, summary: dict):
        """Generate comprehensive Excel report with charts"""
        logger.info("Generating Excel report...")
        
        filename = f"revenue_report_{self.report_timestamp}.xlsx"
        filepath = self.reports_dir / filename
        
        with pd.ExcelWriter(str(filepath), engine='openpyxl') as writer:
            # Executive Summary sheet
            summary_df = pd.DataFrame([
                ['Report Date', summary['report_date']],
                ['Reporting Period', summary['period']],
                ['', ''],
                ['Total Revenue', f"${summary['key_metrics']['total_revenue']:,.0f}"],
                ['YoY Growth', f"{summary['key_metrics']['revenue_growth_yoy']:.1%}" if pd.notna(summary['key_metrics']['revenue_growth_yoy']) else "N/A"],
                ['MoM Growth', f"{summary['key_metrics']['revenue_growth_mom']:.1%}" if pd.notna(summary['key_metrics']['revenue_growth_mom']) else "N/A"],
                ['Total Customers', f"{summary['key_metrics']['total_customers']:,.0f}"],
                ['Revenue per Customer', f"${summary['key_metrics']['avg_revenue_per_customer']:,.0f}"],
                ['Profit Margin', f"{summary['key_metrics']['profit_margin']:.1%}"],
                ['Marketing ROI', f"{summary['key_metrics']['marketing_roi']:.1f}x"],
            ], columns=['Metric', 'Value'])
            
            summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Revenue data sheet
            data['revenue'].to_excel(writer, sheet_name='Revenue Data', index=False)
            
            # KPIs sheet
            data['monthly_kpis'].to_excel(writer, sheet_name='Monthly KPIs', index=False)
            
            # Unit performance sheet
            data['unit_kpis'].to_excel(writer, sheet_name='Unit Performance', index=False)
            
            # Forecasts sheet
            data['forecasts'].to_excel(writer, sheet_name='Revenue Forecasts', index=False)
            
            # Alerts sheet
            if not data['alerts'].empty:
                data['alerts'].to_excel(writer, sheet_name='Performance Alerts', index=False)
            
            # Format the workbook
            self._format_excel_workbook(writer)
        
        logger.info(f"Excel report saved to {filepath}")
    
    def _format_excel_workbook(self, writer):
        """Apply formatting to Excel workbook"""
        workbook = writer.book
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Format each sheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Format headers
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_powerbi_dataset(self, data: dict):
        """Create Power BI compatible datasets"""
        logger.info("Creating Power BI datasets...")
        
        # Create Power BI directory
        powerbi_dir = self.reports_dir / 'powerbi_data'
        powerbi_dir.mkdir(exist_ok=True)
        
        # Save datasets in Power BI friendly format
        datasets = {
            'revenue_data': data['revenue'],
            'kpi_metrics': data['monthly_kpis'],
            'unit_performance': data['unit_kpis'],
            'revenue_forecasts': data['forecasts'],
            'performance_alerts': data['alerts'] if not data['alerts'].empty else pd.DataFrame()
        }
        
        for name, df in datasets.items():
            if not df.empty:
                # Clean column names for Power BI
                df_clean = df.copy()
                df_clean.columns = [col.replace('_', ' ').title() for col in df_clean.columns]
                
                # Save as CSV for Power BI import
                df_clean.to_csv(powerbi_dir / f'{name}.csv', index=False)
        
        # Create Power BI connection script
        connection_script = """
# Power BI Data Connection Script
# Use this script to connect Power BI to the generated datasets

$DataPath = "{data_path}"

# Revenue Data
$RevenueData = Import-Csv "$DataPath/revenue_data.csv"

# KPI Metrics  
$KPIData = Import-Csv "$DataPath/kpi_metrics.csv"

# Unit Performance
$UnitData = Import-Csv "$DataPath/unit_performance.csv"

# Forecasts
$ForecastData = Import-Csv "$DataPath/revenue_forecasts.csv"

# Alerts
$AlertData = Import-Csv "$DataPath/performance_alerts.csv"

Write-Host "Data loaded successfully for Power BI import"
        """.format(data_path=str(powerbi_dir.absolute()).replace('\\', '/'))
        
        with open(powerbi_dir / 'powerbi_connection.ps1', 'w') as f:
            f.write(connection_script)
        
        logger.info(f"Power BI datasets saved to {powerbi_dir}")
    
    def create_summary_dashboard(self, data: dict, summary: dict):
        """Create a visual summary dashboard"""
        logger.info("Creating summary dashboard...")
        
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))
        fig.suptitle(f'{REPORT_TITLE} - {summary["report_date"]}', fontsize=16, fontweight='bold')
        
        # 1. Revenue Trend
        ax1 = axes[0, 0]
        monthly_data = data['monthly_kpis']
        ax1.plot(monthly_data['date'], monthly_data['revenue'] / 1000, linewidth=3, color='blue', marker='o')
        ax1.set_title('Monthly Revenue Trend', fontweight='bold', fontsize=14)
        ax1.set_ylabel('Revenue ($000s)')
        ax1.grid(True, alpha=0.3)
        ax1.tick_params(axis='x', rotation=45)
        
        # Add forecast line
        if not data['forecasts'].empty:
            forecast_by_date = data['forecasts'].groupby('date')['ensemble_forecast'].sum().reset_index()
            ax1.plot(forecast_by_date['date'], forecast_by_date['ensemble_forecast'] / 1000, 
                    linewidth=3, color='red', linestyle='--', marker='s', label='Forecast')
            ax1.legend()
        
        # 2. Business Unit Performance
        ax2 = axes[0, 1]
        unit_data = data['unit_kpis'].sort_values('current_revenue', ascending=True)
        bars = ax2.barh(unit_data['business_unit'], unit_data['current_revenue'] / 1000, color='skyblue')
        ax2.set_title('Revenue by Business Unit', fontweight='bold', fontsize=14)
        ax2.set_xlabel('Revenue ($000s)')
        
        # Add value labels on bars
        for i, bar in enumerate(bars):
            width = bar.get_width()
            ax2.text(width + width*0.02, bar.get_y() + bar.get_height()/2, 
                    f'${width:.0f}K', ha='left', va='center', fontweight='bold')
        
        # 3. YoY Growth by Unit
        ax3 = axes[1, 0]
        growth_data = data['unit_kpis'].dropna(subset=['yoy_growth'])
        colors = ['green' if x >= 0 else 'red' for x in growth_data['yoy_growth']]
        bars = ax3.bar(growth_data['business_unit'], growth_data['yoy_growth'] * 100, color=colors)
        ax3.set_title('Year-over-Year Growth by Unit', fontweight='bold', fontsize=14)
        ax3.set_ylabel('Growth Rate (%)')
        ax3.tick_params(axis='x', rotation=45)
        ax3.axhline(y=REVENUE_GROWTH_TARGET * 100, color='orange', linestyle='--', label='Target')
        ax3.legend()
        
        # Add value labels on bars
        for i, bar in enumerate(bars):
            height = bar.get_height()
            ax3.text(bar.get_x() + bar.get_width()/2, height + (1 if height >= 0 else -3), 
                    f'{height:.1f}%', ha='center', va='bottom' if height >= 0 else 'top', fontweight='bold')
        
        # 4. Alert Summary
        ax4 = axes[1, 1]
        if not data['alerts'].empty:
            alert_counts = data['alerts']['severity'].value_counts()
            colors_alert = {'High': 'red', 'Medium': 'orange', 'Low': 'yellow'}
            pie_colors = [colors_alert.get(x, 'gray') for x in alert_counts.index]
            
            wedges, texts, autotexts = ax4.pie(alert_counts.values, labels=alert_counts.index, 
                                             autopct='%1.0f', colors=pie_colors, startangle=90)
            ax4.set_title('Performance Alerts by Severity', fontweight='bold', fontsize=14)
        else:
            ax4.text(0.5, 0.5, 'No Alerts', ha='center', va='center', fontsize=16, 
                    transform=ax4.transAxes, bbox=dict(boxstyle="round,pad=0.3", facecolor="lightgreen"))
            ax4.set_title('Performance Alerts', fontweight='bold', fontsize=14)
            ax4.axis('off')
        
        plt.tight_layout()
        
        # Save dashboard
        dashboard_path = self.reports_dir / f'executive_dashboard_{self.report_timestamp}.png'
        plt.savefig(dashboard_path, dpi=300, bbox_inches='tight')
        logger.info(f"Executive dashboard saved to {dashboard_path}")
        plt.close()
    
    def generate_all_reports(self):
        """Generate all report formats"""
        logger.info("Starting comprehensive report generation...")
        
        try:
            # Load data
            data = self.load_processed_data()
            
            # Create executive summary
            summary = self.create_executive_summary(data)
            
            # Generate reports
            self.generate_pdf_report(data, summary)
            self.generate_excel_report(data, summary)
            self.create_powerbi_dataset(data)
            self.create_summary_dashboard(data, summary)
            
            # Generate report index
            self._create_report_index(summary)
            
            logger.info("All reports generated successfully!")
            return True
            
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            raise
    
    def _create_report_index(self, summary: dict):
        """Create an index of all generated reports"""
        index_content = f"""
# Revenue Forecasting & Reporting System
## Generated Reports - {summary['report_date']}

### Executive Summary
- **Reporting Period**: {summary['period']}
- **Total Revenue**: ${summary['key_metrics']['total_revenue']:,.0f}
- **YoY Growth**: {summary['key_metrics']['revenue_growth_yoy']:.1%} (if available)
- **Performance Alerts**: {summary['alerts_count']} ({summary['high_priority_alerts']} high priority)

### Generated Files
- **PDF Report**: `revenue_report_{self.report_timestamp}.pdf`
- **Excel Workbook**: `revenue_report_{self.report_timestamp}.xlsx`
- **Executive Dashboard**: `executive_dashboard_{self.report_timestamp}.png`
- **Power BI Data**: `powerbi_data/` directory

### Visualizations
- **Revenue Forecasts**: `revenue_forecasts.png`
- **KPI Dashboard**: `kpi_dashboard.png`

### Data Files
All processed data files are available in the `data/processed/` directory.

---
*Report generated automatically by the Revenue Forecasting & Reporting System*
        """
        
        index_path = self.reports_dir / f'report_index_{self.report_timestamp}.md'
        with open(index_path, 'w') as f:
            f.write(index_content)
        
        logger.info(f"Report index created: {index_path}")

def main():
    """Main execution function"""
    generator = ReportGenerator()
    generator.generate_all_reports()

if __name__ == "__main__":
    main()